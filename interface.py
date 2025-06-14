import streamlit as st
import numpy as np
import pandas as pd
import plotly.express as px
import requests
from datetime import datetime
import pytz
from ahp_calculator import calculate_ahp
import openpyxl
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Hàm gọi API
def safe_api_request(method, endpoint, data=None):
    base_url = "http://127.0.0.1:8000"
    url = f"{base_url}/{endpoint}"
    try:
        if method.upper() == "GET":
            response = requests.get(url)
        elif method.upper() == "POST":
            print(f"Sending POST to {url} with data: {data}")  # Debug dữ liệu gửi
            response = requests.post(url, json=data)
        elif method.upper() == "DELETE":
            response = requests.delete(url)
        else:
            return {"error": "Phương thức HTTP không hỗ trợ"}
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        return {"error": str(e), "status_code": getattr(e.response, "status_code", None)}

# Hàm hiển thị ma trận dưới dạng HTML
def matrix_to_html(matrix, labels):
    html = "<table border='1'>"
    html += "<tr><th></th>" + "".join(f"<th>{label}</th>" for label in labels) + "</tr>"
    for i, row in enumerate(matrix):
        html += f"<tr><td>{labels[i]}</td>" + "".join(f"<td>{val:.2f}</td>" for val in row) + "</tr>"
    html += "</table>"
    return html

# Hàm kiểm tra trạng thái
def check_status():
    response = safe_api_request("GET", "check_status")
    if "error" in response:
        return {"criteria_count": 0, "vehicles_count": 0, "weights_saved": False}
    return response

# Hàm chuyển đổi giá trị thành số thực
def convert_to_float(value):
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.strip()
        if '/' in value:
            try:
                num, denom = map(float, value.split('/'))
                if denom == 0:
                    raise ValueError("Mẫu số không được bằng 0")
                return num / denom
            except (ValueError, ZeroDivisionError):
                raise ValueError(f"Phân số không hợp lệ: {value}")
        try:
            return float(value)
        except ValueError:
            raise ValueError(f"Giá trị không phải số: {value}")
    raise ValueError(f"Giá trị không hợp lệ: {value}")

# Hàm tạo file Excel từ dữ liệu AHP
def create_excel_file(criteria_names, vehicle_names, criteria_matrix, alternative_matrices, criteria_weights, ranking, consistency_ratios, criteria_cr):
    output = BytesIO()
    wb = openpyxl.Workbook()
    
    # Sheet MTSS (ma trận tiêu chí)
    ws = wb.active
    ws.title = "MTSS"
    ws.append([""] + criteria_names)
    for i, row in enumerate(criteria_matrix):
        ws.append([criteria_names[i]] + [f"{val:.2f}" if val != int(val) else str(int(val)) for val in row])
    
    # Sheets MTSS - <tên tiêu chí> (ma trận xe)
    for crit_idx, criterion in enumerate(criteria_names):
        ws = wb.create_sheet(f"MTSS - {criterion}")
        ws.append([""] + vehicle_names)
        matrix = alternative_matrices[crit_idx]
        for i, row in enumerate(matrix):
            ws.append([vehicle_names[i]] + [f"{val:.2f}" if val != int(val) else str(int(val)) for val in row])
    
    # Sheet Results
    ws = wb.create_sheet("Results")
    ws.append(["Trọng số tiêu chí"])
    for name, weight in zip(criteria_names, criteria_weights):
        ws.append([name, f"{weight:.4f}"])
    ws.append([])
    ws.append(["Xếp hạng xe"])
    ws.append(["Xe", "Điểm AHP"])
    for name, score in ranking:
        ws.append([name, f"{score:.4f}"])
    ws.append([])
    ws.append(["Độ nhất quán"])
    ws.append(["Tiêu chí", "CR"])
    ws.append(["Tiêu chí tổng thể", f"{criteria_cr:.4f}"])
    for name, cr in zip(criteria_names, consistency_ratios):
        ws.append([name, f"{cr:.4f}"])
    
    wb.save(output)
    return output.getvalue()

# Hàm tạo file PDF từ dữ liệu AHP
def create_pdf_file(criteria_names, vehicle_names, criteria_matrix, alternative_matrices, criteria_weights, ranking, consistency_ratios, criteria_cr):
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=0.75*inch, leftMargin=0.75*inch, topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()

    # Đăng ký font Arial hỗ trợ tiếng Việt
    try:
        pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
    except Exception as e:
        # Nếu không tìm thấy arial.ttf, sử dụng font mặc định và cảnh báo
        st.warning("Font Arial không khả dụng, sử dụng font mặc định. Có thể không hiển thị đúng tiếng Việt.")
        print(f"Lỗi khi đăng ký font Arial: {str(e)}")

    # Cập nhật style để sử dụng font Arial
    title_style = ParagraphStyle(name='Title', fontName='Arial', fontSize=16, alignment=1, spaceAfter=12)
    heading_style = ParagraphStyle(name='Heading2', fontName='Arial', fontSize=12, spaceAfter=12)
    normal_style = ParagraphStyle(name='Normal', fontName='Arial', fontSize=10, spaceAfter=12)

    elements = []

    # Tiêu đề
    elements.append(Paragraph("Kết quả phân tích AHP", title_style))
    elements.append(Spacer(1, 12))

    # Ma trận tiêu chí
    elements.append(Paragraph("Ma trận so sánh cặp tiêu chí", heading_style))
    data = [[""] + criteria_names]
    for i, row in enumerate(criteria_matrix):
        data.append([criteria_names[i]] + [f"{val:.2f}" if val != int(val) else str(int(val)) for val in row])
    table = Table(data)
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Arial'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    elements.append(Paragraph(f"Độ nhất quán (CR): {criteria_cr:.4f} ({'Nhất quán' if criteria_cr < 0.1 else 'Không nhất quán'})", normal_style))
    elements.append(Spacer(1, 12))

    # Ma trận xe
    for crit_idx, criterion in enumerate(criteria_names):
        elements.append(Paragraph(f"Ma trận so sánh cặp xe - {criterion}", heading_style))
        data = [[""] + vehicle_names]
        matrix = alternative_matrices[crit_idx]
        for i, row in enumerate(matrix):
            data.append([vehicle_names[i]] + [f"{val:.2f}" if val != int(val) else str(int(val)) for val in row])
        table = Table(data)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), 'Arial'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)
        elements.append(Paragraph(f"Độ nhất quán (CR): {consistency_ratios[crit_idx]:.4f} ({'Nhất quán' if consistency_ratios[crit_idx] < 0.1 else 'Không nhất quán'})", normal_style))
        elements.append(Spacer(1, 12))

    # Trọng số tiêu chí
    elements.append(Paragraph("Trọng số tiêu chí", heading_style))
    data = [["Tiêu chí", "Trọng số"]]
    for name, weight in zip(criteria_names, criteria_weights):
        data.append([name, f"{weight:.4f}"])
    table = Table(data)
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Arial'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12))

    # Kết quả xếp hạng
    elements.append(Paragraph("Xếp hạng xe", heading_style))
    data = [["Xe", "Điểm AHP"]]
    for name, score in ranking:
        data.append([name, f"{score:.4f}"])
    table = Table(data)
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Arial'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12))

    # Độ nhất quán
    elements.append(Paragraph("Độ nhất quán", heading_style))
    data = [["Tiêu chí", "CR"]]
    data.append(["Tiêu chí tổng thể", f"{criteria_cr:.4f}"])
    for name, cr in zip(criteria_names, consistency_ratios):
        data.append([name, f"{cr:.4f}"])
    table = Table(data)
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Arial'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)

    doc.build(elements)
    return output.getvalue()

# Bước 1: Quản lý tiêu chí và xe
def criteria_management_step():
    st.markdown("## Bước 1: Quản lý tiêu chí và xe")
    st.markdown("Thêm ít nhất 2 tiêu chí và 2 xe để tiếp tục.")

    def fetch_criteria():
        response = safe_api_request("GET", "get_criteria")
        return response.get("criteria", []) if "error" not in response else []

    def fetch_vehicles():
        response = safe_api_request("GET", "get_vehicles")
        return response.get("vehicles", []) if "error" not in response else []

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("### Quản lý tiêu chí")
        new_criterion = st.text_input("Tên tiêu chí mới")
        if st.button("Thêm tiêu chí"):
            if new_criterion:
                response = safe_api_request("POST", "add_criterion", {"name": new_criterion})
                if "error" in response:
                    st.error(f"Lỗi: {response['error']} (Mã trạng thái: {response.get('status_code', 'N/A')})")
                else:
                    st.success(response.get("message", "Thêm tiêu chí thành công."))
                    st.rerun()
            else:
                st.error("Vui lòng nhập tên tiêu chí.")

        criteria = fetch_criteria()
        criterion_to_delete = st.selectbox("Chọn tiêu chí để xóa", [c["name"] for c in criteria], index=None)
        if st.button("Xóa tiêu chí"):
            if criterion_to_delete:
                response = safe_api_request("DELETE", f"delete_criterion/{criterion_to_delete}")
                if "error" in response:
                    st.error(f"Lỗi: {response['error']} (Mã trạng thái: {response.get('status_code', 'N/A')})")
                else:
                    st.success(response.get("message", "Xóa tiêu chí thành công."))
                    st.rerun()
            else:
                st.error("Vui lòng chọn tiêu chí để xóa.")

        st.markdown("#### Danh sách tiêu chí:")
        for c in criteria:
            st.write(f"- {c['name']} (index: {c['index']})")

    with col2:
        st.markdown("### Quản lý xe")
        vehicle_name = st.text_input("Tên xe")
        vehicle_image = st.text_input("URL hình ảnh", value="")
        vehicle_brand = st.text_input("Hãng xe", value="N/A")
        vehicle_year = st.number_input("Năm sản xuất", min_value=1900, max_value=2025, value=2020)
        vehicle_type = st.text_input("Loại xe", value="N/A")
        vehicle_engine = st.text_input("Động cơ", value="N/A")
        vehicle_price = st.text_input("Khoảng giá", value="N/A")
        if st.button("Thêm xe"):
            if vehicle_name:
                vehicle_data = {
                    "name": vehicle_name,
                    "image": vehicle_image,
                    "details": {
                        "brand": vehicle_brand,
                        "year": float(vehicle_year),
                        "type": vehicle_type,
                        "engine": vehicle_engine,
                        "price_range": vehicle_price
                    }
                }
                response = safe_api_request("POST", "add_vehicle", vehicle_data)
                if "error" in response:
                    st.error(f"Lỗi: {response['error']} (Mã trạng thái: {response.get('status_code', 'N/A')})")
                else:
                    st.success(response.get("message", "Thêm xe thành công."))
                    st.rerun()
            else:
                st.error("Vui lòng nhập tên xe.")

        vehicles = fetch_vehicles()
        vehicle_to_delete = st.selectbox("Chọn xe để xóa", [v["name"] for v in vehicles], index=None)
        if st.button("Xóa xe"):
            if vehicle_to_delete:
                response = safe_api_request("DELETE", f"delete_vehicle/{vehicle_to_delete}")
                if "error" in response:
                    st.error(f"Lỗi: {response['error']} (Mã trạng thái: {response.get('status_code', 'N/A')})")
                else:
                    st.success(response.get("message", "Xóa xe thành công."))
                    st.rerun()
            else:
                st.error("Vui lòng chọn xe để xóa.")

        st.markdown("#### Danh sách xe:")
        for v in vehicles:
            details = v.get("details", {})
            st.write(f"- {v['name']} ({details.get('brand', 'N/A')}, {details.get('year', 'N/A')})")

    status = check_status()
    criteria_count = status.get("criteria_count", 0)
    vehicles_count = status.get("vehicles_count", 0)
    if criteria_count < 2 or vehicles_count < 2:
        st.error(f"Cần ít nhất 2 tiêu chí (hiện có: {criteria_count}) và 2 xe (hiện có: {vehicles_count}).")
    else:
        st.success("Đủ tiêu chí và xe. Nhấn 'Tiếp tục' để sang bước tiếp theo.")
        if st.button("Tiếp tục đến bước So sánh cặp tiêu chí"):
            st.session_state.step = "pairwise_comparison"
            st.rerun()

# Bước 2: So sánh cặp tiêu chí
def pairwise_comparison_step():
    st.markdown("## Bước 2: So sánh cặp các tiêu chí")
    st.markdown("""
    ### Thang đánh giá từ 1 đến 9:
    - **1**: Hai tiêu chí có tầm quan trọng ngang nhau
    - **3**: Tiêu chí A quan trọng hơn tiêu chí B một chút
    - **5**: Tiêu chí A quan trọng hơn tiêu chí B nhiều
    - **7**: Tiêu chí A quan trọng hơn tiêu chí B rất nhiều
    - **9**: Tiêu chí A quan trọng hơn tiêu chí B cực kỳ nhiều
    - **2, 4, 6, 8**: Giá trị trung gian
    """)

    def fetch_criteria():
        response = safe_api_request("GET", "get_criteria")
        return response.get("criteria", []) if "error" not in response else []

    criteria = fetch_criteria()
    criteria_names = [c["name"] for c in criteria]
    if len(criteria_names) < 2:
        st.error("Cần ít nhất 2 tiêu chí để so sánh. Vui lòng quay lại bước 1.")
        if st.button("Quay lại bước 1"):
            st.session_state.step = "criteria_management"
            st.rerun()
        return

    inputs = []
    for i in range(len(criteria_names)):
        for j in range(i + 1, len(criteria_names)):
            key = f"criteria_{criteria_names[i]}_vs_{criteria_names[j]}"
            value = st.session_state.get(key, 1.0)
            slider = st.slider(
                f"{criteria_names[i]} so với {criteria_names[j]}",
                min_value=0.1, max_value=9.0, step=0.1, value=value,
                key=key
            )
            inputs.append(slider)

    def update_matrix_and_weights(inputs):
        n = len(criteria_names)
        matrix = np.ones((n, n))
        input_idx = 0
        for i in range(n):
            for j in range(i + 1, n):
                matrix[i][j] = inputs[input_idx]
                matrix[j][i] = 1.0 / inputs[input_idx]
                input_idx += 1

        col_sums = np.sum(matrix, axis=0)
        normalized_matrix = matrix / col_sums
        weights = np.mean(normalized_matrix, axis=1)
        weighted_sum = np.dot(matrix, weights)
        consistency_vector = weighted_sum / weights
        lambda_max = np.mean(consistency_vector)
        ci = (lambda_max - n) / (n - 1) if n > 1 else 0
        ri_values = {1: 0, 2: 0, 3: 0.58, 4: 0.9, 5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49}
        ri = ri_values.get(n, 1.5)
        cr = ci / ri if ri != 0 else 0

        matrix_html = matrix_to_html(matrix, criteria_names)
        weights_html = "<h4>Trọng số:</h4><ul>"
        for i, w in enumerate(weights):
            weights_html += f"<li>{criteria_names[i]}: {w:.4f}</li>"
        weights_html += f"</ul><p>CR = {cr:.4f} ({'Nhất quán' if cr < 0.1 else 'Không nhất quán'})</p>"

        return matrix_html, weights_html, weights, cr, matrix

    matrix_html, weights_html, weights, cr, criteria_matrix = update_matrix_and_weights(inputs)
    st.markdown("### Ma trận so sánh cặp")
    st.markdown(matrix_html, unsafe_allow_html=True)
    st.markdown("### Trọng số tiêu chí")
    st.markdown(weights_html, unsafe_allow_html=True)

    if st.button("Lưu trọng số tiêu chí"):
        if cr >= 0.1:
            st.error(f"Ma trận không nhất quán (CR = {cr:.4f}). Vui lòng điều chỉnh giá trị so sánh.")
        else:
            response = safe_api_request("POST", "save_criteria_weights", {"weights": weights.tolist()})
            if "error" in response:
                st.error(f"Lỗi: {response['error']} (Mã trạng thái: {response.get('status_code', 'N/A')})")
            else:
                st.success(response.get("message", "Lưu trọng số thành công."))
                st.session_state.weights_saved = True
                st.session_state.criteria_matrix = criteria_matrix
                st.rerun()

    if st.session_state.get("weights_saved", False):
        if st.button("Tiếp tục đến bước So sánh cặp xe"):
            st.session_state.step = "alternatives_comparison"
            st.rerun()
    else:
        st.error("Vui lòng lưu trọng số trước khi tiếp tục.")

    if st.button("Quay lại bước 1"):
        st.session_state.step = "criteria_management"
        st.rerun()

# Bước 3: So sánh cặp xe
def alternatives_comparison_step():
    st.markdown("## Bước 3: So sánh cặp các xe theo từng tiêu chí")
    st.markdown("Nhập giá trị so sánh cặp (1-9) cho tất cả xe theo từng tiêu chí.")

    def fetch_criteria():
        response = safe_api_request("GET", "get_criteria")
        return response.get("criteria", []) if "error" not in response else []

    def fetch_vehicles():
        response = safe_api_request("GET", "get_vehicles")
        return response.get("vehicles", []) if "error" not in response else []

    vehicles = fetch_vehicles()
    vehicle_names = [v["name"] for v in vehicles]
    criteria = fetch_criteria()
    criteria_names = [c["name"] for c in criteria]

    if len(criteria_names) == 0:
        st.error("Không có tiêu chí nào. Vui lòng quay lại bước 1.")
        if st.button("Quay lại bước 1"):
            st.session_state.step = "criteria_management"
            st.rerun()
        return
    if len(vehicle_names) < 2:
        st.error("Cần ít nhất 2 xe. Vui lòng quay lại bước 1.")
        if st.button("Quay lại bước 1"):
            st.session_state.step = "criteria_management"
            st.rerun()
        return

    alternative_inputs = []
    matrix_outputs = []

    tabs = st.tabs(criteria_names)

    for crit_idx, tab in enumerate(tabs):
        with tab:
            st.markdown(f"### Tiêu chí: {criteria_names[crit_idx]}")
            inputs = []
            for i in range(len(vehicle_names)):
                for j in range(i + 1, len(vehicle_names)):
                    key = f"{criteria_names[crit_idx]}_{vehicle_names[i]}_vs_{vehicle_names[j]}"
                    value = st.session_state.get(key, 1.0)
                    slider = st.slider(
                        f"{vehicle_names[i]} so với {vehicle_names[j]}",
                        min_value=0.1, max_value=9.0, step=0.1, value=value,
                        key=key
                    )
                    inputs.append(slider)
            alternative_inputs.append(inputs)

            matrix = np.ones((len(vehicle_names), len(vehicle_names)))
            input_idx = 0
            for i in range(len(vehicle_names)):
                for j in range(i + 1, len(vehicle_names)):
                    matrix[i][j] = inputs[input_idx]
                    matrix[j][i] = 1.0 / inputs[input_idx]
                    input_idx += 1
            col_sums = np.sum(matrix, axis=0)
            normalized_matrix = matrix / col_sums
            weights = np.mean(normalized_matrix, axis=1)
            weighted_sum = np.dot(matrix, weights)
            consistency_vector = weighted_sum / weights
            lambda_max = np.mean(consistency_vector)
            ci = (lambda_max - len(vehicle_names)) / (len(vehicle_names) - 1) if len(vehicle_names) > 1 else 0
            ri_values = {1: 0, 2: 0, 3: 0.58, 4: 0.9, 5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49}
            ri = ri_values.get(len(vehicle_names), 1.5)
            cr = ci / ri if ri != 0 else 0
            html = matrix_to_html(matrix, vehicle_names)
            html += f"<p>CR = {cr:.4f} ({'Nhất quán' if cr < 0.1 else 'Không nhất quán'})</p>"
            matrix_outputs.append(html)

            st.markdown(f"#### Ma trận so sánh cặp ({criteria_names[crit_idx]})")
            st.markdown(html, unsafe_allow_html=True)

    if st.button("Tính toán AHP"):
        response = safe_api_request("GET", "get_criteria_weights")
        if "error" in response:
            st.error(f"Lỗi: {response['error']} (Mã trạng thái: {response.get('status_code', 'N/A')})")
            return
        criteria_weights = response.get("weights", [1/len(criteria_names)] * len(criteria_names))

        alternative_matrices = []
        consistency_ratios = []
        ri_values = {1: 0, 2: 0, 3: 0.58, 4: 0.9, 5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49}
        for crit_idx in range(len(criteria_names)):
            matrix = np.ones((len(vehicle_names), len(vehicle_names)))
            inputs = alternative_inputs[crit_idx]
            input_idx = 0
            for i in range(len(vehicle_names)):
                for j in range(i + 1, len(vehicle_names)):
                    matrix[i][j] = inputs[input_idx]
                    matrix[j][i] = 1.0 / inputs[input_idx]
                    input_idx += 1
            col_sums = np.sum(matrix, axis=0)
            normalized_matrix = matrix / col_sums
            weights = np.mean(normalized_matrix, axis=1)
            weighted_sum = np.dot(matrix, weights)
            consistency_vector = weighted_sum / weights
            lambda_max = np.mean(consistency_vector)
            ci = (lambda_max - len(vehicle_names)) / (len(vehicle_names) - 1) if len(vehicle_names) > 1 else 0
            ri = ri_values.get(len(vehicle_names), 1.5)
            cr = ci / ri if ri != 0 else 0
            consistency_ratios.append(cr)
            alternative_matrices.append(matrix)

        try:
            ranking = calculate_ahp(criteria_weights, alternative_matrices, vehicle_names)
        except Exception as e:
            st.error(f"Lỗi khi tính AHP: {str(e)}")
            return

        inconsistent_matrices = [i for i, cr in enumerate(consistency_ratios) if cr >= 0.1]
        warning = f"Cảnh báo: Ma trận không nhất quán (CR ≥ 0.1) cho tiêu chí: {[criteria_names[i] for i in inconsistent_matrices]}" if inconsistent_matrices else None

        html_result = "<h3>Kết quả xếp hạng:</h3>"
        if warning:
            html_result += f"<p style='color:red'>{warning}</p>"
        html_result += "<table border='1'><tr><th>Xe</th><th>Điểm AHP</th></tr>"
        for name, score in ranking:
            html_result += f"<tr><td>{name}</td><td>{score:.4f}</td></tr>"
        html_result += "</table>"
        html_result += "<h3>Độ nhất quán:</h3><ul>"
        for i, cr in enumerate(consistency_ratios):
            html_result += f"<li>Tiêu chí {criteria_names[i]}: CR = {cr:.4f} ({'Nhất quán' if cr < 0.1 else 'Không nhất quán'})</li>"
        html_result += "</ul>"

        st.markdown(html_result, unsafe_allow_html=True)

        fig = px.bar(x=[name for name, score in ranking], y=[score for name, score in ranking],
                     labels={'x': 'Xe', 'y': 'Điểm AHP'}, title='Xếp hạng xe')
        fig.update_traces(texttemplate='%{y:.4f}', textposition='outside')
        fig.update_layout(xaxis={'categoryorder': 'total descending'})
        st.plotly_chart(fig)

        weights_fig = px.pie(values=criteria_weights, names=criteria_names,
                            title='Trọng số tiêu chí', hole=0.4)
        weights_fig.update_traces(textinfo='percent+label')
        st.plotly_chart(weights_fig)

        log_data = {
            "weights": criteria_weights,
            "top_result": [[name, score] for name, score in ranking[:3]],
            "criteria_matrices": [{"vehicle_" + str(i+1) + "_vs_" + str(j+1): matrix[i][j]
                                  for i in range(len(vehicle_names))
                                  for j in range(i + 1, len(vehicle_names))}
                                 for matrix in alternative_matrices]
        }
        response = safe_api_request("POST", "log_calculation", log_data)
        if "error" in response:
            st.error(f"Lỗi khi lưu log: {response['error']} (Mã trạng thái: {response.get('status_code', 'N/A')})")
        else:
            st.success("Lưu log tính toán thành công.")

        criteria_matrix = st.session_state.get("criteria_matrix", np.ones((len(criteria_names), len(criteria_names))))
        response = safe_api_request("GET", "get_criteria_weights")
        criteria_weights = response.get("weights", [1/len(criteria_names)] * len(criteria_names))
        criteria_cr_response = safe_api_request("GET", "get_criteria_weights")
        criteria_cr = criteria_cr_response.get("cr", 0.0) if "cr" in criteria_cr_response else 0.0

        excel_data = create_excel_file(
            criteria_names, vehicle_names, criteria_matrix, alternative_matrices,
            criteria_weights, ranking, consistency_ratios, criteria_cr
        )
        st.download_button(
            label="Xuất kết quả Excel",
            data=excel_data,
            file_name="ahp_results_manual.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        pdf_data = create_pdf_file(
            criteria_names, vehicle_names, criteria_matrix, alternative_matrices,
            criteria_weights, ranking, consistency_ratios, criteria_cr
        )
        st.download_button(
            label="Xuất kết quả PDF",
            data=pdf_data,
            file_name="ahp_results_manual.pdf",
            mime="application/pdf"
        )

    if st.button("Quay lại bước 2"):
        st.session_state.step = "pairwise_comparison"
        st.rerun()

# Bước 4: Tính AHP từ Excel
def excel_calculation_step():
    st.markdown("## Bước 4: Tính AHP từ file Excel")
    st.markdown("""
    Tải lên file Excel chứa ma trận so sánh tiêu chí (sheet 'MTSS') và ma trận so sánh xe (sheets 'MTSS - {tên tiêu chí}').
    File cần có cấu trúc như sau:
    - **Sheet `MTSS`**: Ma trận so sánh cặp giữa các tiêu chí, với tiêu chí ở cột và hàng đầu tiên.
    - **Sheets `MTSS - <tên tiêu chí>`**: Ma trận so sánh cặp giữa các xe cho từng tiêu chí, với tên xe ở cột và hàng đầu tiên.
    - Giá trị so sánh có thể là số nguyên (1, 3), thập phân (0.5), hoặc phân số (1/2, 1/3).
    - Ô trên đường chéo chính phải là 1. Không để trống ô.
    """)

    with st.expander("Xem ví dụ định dạng file Excel"):
        st.markdown("""
        ### Ví dụ file Excel với 3 tiêu chí (Độ bền, Hiệu suất, Giá bán) và 3 xe (Winner X, NVX 155, Vario 160)
        
        #### Sheet `MTSS` (Ma trận so sánh tiêu chí)
        |               | Độ bền | Hiệu suất | Giá bán |
        |---------------|--------|-----------|---------|
        | **Độ bền**    | 1      | 3         | 5       |
        | **Hiệu suất** | 1/3    | 1         | 2       |
        | **Giá bán**   | 1/5    | 1/2       | 1       |

        #### Sheet `MTSS - Độ bền` (Ma trận so sánh xe)
        |               | Winner X | NVX 155 | Vario 160 |
        |---------------|----------|---------|-----------|
        | **Winner X**  | 1        | 2       | 4         |
        | **NVX 155**   | 1/2      | 1       | 2         |
        | **Vario 160** | 1/4      | 1/2     | 1         |

        #### Sheet `MTSS - Hiệu suất`
        |               | Winner X | NVX 155 | Vario 160 |
        |---------------|----------|---------|-----------|
        | **Winner X**  | 1        | 1/2     | 1/3       |
        | **NVX 155**   | 2        | 1       | 1/2       |
        | **Vario 160** | 3        | 2       | 1         |

        #### Sheet `MTSS - Giá bán`
        |               | Winner X | NVX 155 | Vario 160 |
        |---------------|----------|---------|-----------|
        | **Winner X**  | 1        | 3       | 5         |
        | **NVX 155**   | 1/3      | 1       | 2         |
        | **Vario 160** | 1/5      | 1/2     | 1         |
        """)

    def compute_weights_and_cr(matrix, n):
        col_sums = np.sum(matrix, axis=0)
        if np.any(col_sums == 0):
            raise ValueError("Ma trận chứa cột có tổng bằng 0")
        normalized_matrix = matrix / col_sums
        weights = np.mean(normalized_matrix, axis=1)
        weights = weights / np.sum(weights)
        weighted_sum = np.dot(matrix, weights)
        consistency_vector = weighted_sum / weights
        lambda_max = np.mean(consistency_vector)
        ci = (lambda_max - n) / (n - 1) if n > 1 else 0
        ri_values = {1: 0, 2: 0, 3: 0.58, 4: 0.9, 5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49}
        ri = ri_values.get(n, 1.5)
        cr = ci / ri if ri != 0 else 0
        return weights, cr

    if "excel_log_saved" not in st.session_state or st.session_state.get("uploaded_file_key") != st.session_state.get("excel_uploader"):
        st.session_state.excel_log_saved = False
        st.session_state.uploaded_file_key = st.session_state.get("excel_uploader")

    uploaded_file = st.file_uploader("Chọn file Excel", type=["xlsx"], key="excel_uploader")

    if uploaded_file:
        try:
            xls = pd.ExcelFile(uploaded_file)
            sheets = xls.sheet_names

            if "MTSS" not in sheets:
                st.error("File Excel phải chứa sheet 'MTSS' cho ma trận so sánh tiêu chí.")
                return

            criteria_df = pd.read_excel(uploaded_file, sheet_name="MTSS", index_col=0)
            criteria_names = criteria_df.index.tolist()
            if len(criteria_names) < 2:
                st.error("Cần ít nhất 2 tiêu chí trong sheet 'MTSS'.")
                return

            try:
                criteria_df = criteria_df.applymap(convert_to_float)
            except ValueError as e:
                st.error(f"Lỗi khi chuyển đổi dữ liệu ma trận tiêu chí: {str(e)}")
                return

            criteria_df = criteria_df.fillna(0)
            criteria_matrix = criteria_df.to_numpy(dtype=float)

            if criteria_matrix.shape != (len(criteria_names), len(criteria_names)):
                st.error("Ma trận tiêu chí phải là ma trận vuông.")
                return
            if np.any(np.isnan(criteria_matrix)) or np.any(np.isinf(criteria_matrix)):
                st.error("Ma trận tiêu chí chứa NaN hoặc giá trị vô cực.")
                return

            criteria_weights, criteria_cr = compute_weights_and_cr(criteria_matrix, len(criteria_names))

            st.markdown("### Ma trận so sánh cặp tiêu chí")
            st.markdown(matrix_to_html(criteria_matrix, criteria_names), unsafe_allow_html=True)
            st.markdown(f"### Độ nhất quán tiêu chí: CR = {criteria_cr:.4f} ({'Nhất quán' if criteria_cr < 0.1 else 'Không nhất quán'})")

            vehicle_names = None
            alternative_matrices = []
            consistency_ratios = []
            for criterion in criteria_names:
                sheet_name = f"MTSS - {criterion}"
                if sheet_name not in sheets:
                    st.error(f"File Excel phải chứa sheet '{sheet_name}' cho tiêu chí '{criterion}'.")
                    return
                df = pd.read_excel(uploaded_file, sheet_name=sheet_name, index_col=0)
                if vehicle_names is None:
                    vehicle_names = df.index.tolist()
                    if len(vehicle_names) < 2:
                        st.error("Cần ít nhất 2 xe trong ma trận so sánh.")
                        return
                elif df.index.tolist() != vehicle_names:
                    st.error(f"Sheet '{sheet_name}' có danh sách xe không khớp với các sheet khác.")
                    return

                try:
                    df = df.applymap(convert_to_float)
                except ValueError as e:
                    st.error(f"Lỗi khi chuyển đổi dữ liệu ma trận '{criterion}': {str(e)}")
                    return

                df = df.fillna(0)
                matrix = df.to_numpy(dtype=float)

                if matrix.shape != (len(vehicle_names), len(vehicle_names)):
                    st.error(f"Ma trận cho tiêu chí '{criterion}' phải là ma trận vuông.")
                    return
                if np.any(np.isnan(matrix)) or np.any(np.isinf(matrix)):
                    st.error(f"Ma trận cho tiêu chí '{criterion}' chứa NaN hoặc giá trị vô cực.")
                    return

                _, cr = compute_weights_and_cr(matrix, len(vehicle_names))
                consistency_ratios.append(cr)
                alternative_matrices.append(matrix)

                st.markdown(f"### Ma trận so sánh cặp xe ({criterion})")
                st.markdown(matrix_to_html(matrix, vehicle_names) + f"<p>CR = {cr:.4f} ({'Nhất quán' if cr < 0.1 else 'Không nhất quán'})</p>", unsafe_allow_html=True)

            inconsistent_matrices = []
            if criteria_cr >= 0.1:
                inconsistent_matrices.append("Tiêu chí tổng thể")
            for i, cr in enumerate(consistency_ratios):
                if cr >= 0.1:
                    inconsistent_matrices.append(f"Tiêu chí {criteria_names[i]}")

            if not inconsistent_matrices:
                st.success("Tất cả ma trận đều nhất quán (CR < 0.1). Nhấn 'Tính AHP' để tiếp tục.")
                calculate_enabled = True
            else:
                st.error(f"Các ma trận không nhất quán (CR ≥ 0.1): {', '.join(inconsistent_matrices)}. Vui lòng sửa file Excel.")
                calculate_enabled = False

            if st.button("Tính AHP", disabled=not calculate_enabled):
                n_vehicles = len(vehicle_names)
                scores = np.zeros(n_vehicles)
                for i, matrix in enumerate(alternative_matrices):
                    weights, _ = compute_weights_and_cr(matrix, n_vehicles)
                    scores += criteria_weights[i] * weights

                ranking = [(vehicle_names[i], scores[i]) for i in range(n_vehicles)]
                ranking.sort(key=lambda x: x[1], reverse=True)

                html_result = "<h3>Kết quả xếp hạng:</h3>"
                html_result += "<table border='1'><tr><th>Xe</th><th>Điểm AHP</th></tr>"
                for name, score in ranking:
                    html_result += f"<tr><td>{name}</td><td>{score:.4f}</td></tr>"
                html_result += "</table>"
                st.markdown(html_result, unsafe_allow_html=True)

                fig = px.bar(x=[name for name, score in ranking], y=[score for name, score in ranking],
                             labels={'x': 'Xe', 'y': 'Điểm AHP'}, title='Xếp hạng xe')
                fig.update_traces(texttemplate='%{y:.4f}', textposition='outside')
                fig.update_layout(xaxis={'categoryorder': 'total descending'})
                st.plotly_chart(fig)

                weights_fig = px.pie(values=criteria_weights, names=criteria_names,
                                    title='Trọng số tiêu chí', hole=0.4)
                weights_fig.update_traces(textinfo='percent+label')
                st.plotly_chart(weights_fig)

                log_data = {
                    "weights": criteria_weights.tolist(),
                    "top_result": [[name, score] for name, score in ranking[:3]],
                    "criteria_matrices": [{"vehicle_" + str(i+1) + "_vs_" + str(j+1): matrix[i][j]
                                          for i in range(len(vehicle_names))
                                          for j in range(i + 1, len(vehicle_names))}
                                         for matrix in alternative_matrices]
                }
                response = safe_api_request("POST", "log_calculation", log_data)
                if "error" in response:
                    st.error(f"Lỗi khi lưu log: {response['error']} (Mã trạng thái: {response.get('status_code', 'N/A')})")
                else:
                    st.success("Lưu log tính toán thành công.")
                    st.session_state.excel_log_saved = True

                excel_data = create_excel_file(
                    criteria_names, vehicle_names, criteria_matrix, alternative_matrices,
                    criteria_weights, ranking, consistency_ratios, criteria_cr
                )
                st.download_button(
                    label="Xuất kết quả Excel",
                    data=excel_data,
                    file_name="ahp_results_excel.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                pdf_data = create_pdf_file(
                    criteria_names, vehicle_names, criteria_matrix, alternative_matrices,
                    criteria_weights, ranking, consistency_ratios, criteria_cr
                )
                st.download_button(
                    label="Xuất kết quả PDF",
                    data=pdf_data,
                    file_name="ahp_results_excel.pdf",
                    mime="application/pdf"
                )

        except Exception as e:
            st.error(f"Lỗi khi xử lý file Excel: {str(e)}")
            return

    if st.button("Quay lại bước 1"):
        st.session_state.step = "criteria_management"
        st.rerun()

# Bước 5: Xem lịch sử tính toán
def log_step():
    if "previous_step" not in st.session_state:
        st.session_state.previous_step = st.session_state.get("step", "criteria_management")

    st.markdown("## Xem lịch sử tính toán")

    def fetch_logs():
        response = safe_api_request("GET", "logs")
        if "error" in response:
            return f"<p style='color:red'>Lỗi: {response['error']} (Mã trạng thái: {response.get('status_code', 'N/A')})</p>", []
        logs = response
        if not logs:
            return "<p>Không có log nào.</p>", []
        html = "<h4>Lịch sử tính toán:</h4>"
        choices = []
        for log in logs:
            log_id = str(log.get("_id", ""))
            if not log_id or len(log_id) != 24:
                continue
            timestamp = log.get("timestamp", "N/A")
            try:
                dt = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
                formatted_timestamp = dt.astimezone(pytz.timezone("Asia/Ho_Chi_Minh")).strftime("%d/%m/%Y %H:%M:%S")
            except (ValueError, TypeError):
                formatted_timestamp = timestamp
            weights = log.get("weights", [])
            top_result = log.get("top_result", [])
            html += f"<p><b>Thời gian:</b> {formatted_timestamp}</p>"
            html += "<p><b>Trọng số tiêu chí:</b> " + ", ".join([f"{w:.4f}" for w in weights]) + "</p>"
            html += "<p><b>Kết quả top:</b></p><ul>"
            for name, score in top_result:
                html += f"<li>{name}: {score:.4f}</li>"
            html += "</ul><hr>"
            choices.append((formatted_timestamp, log_id))
        return html, choices

    logs_html, choices = fetch_logs()
    st.markdown(logs_html, unsafe_allow_html=True)

    log_to_delete = st.selectbox("Chọn log để xóa (thời gian)", [c[0] for c in choices], index=None)
    if st.button("Xóa log"):
        if log_to_delete:
            log_id = next(c[1] for c in choices if c[0] == log_to_delete)
            response = safe_api_request("DELETE", f"logs/{log_id}")
            if "error" in response:
                st.error(f"Lỗi: {response['error']} (Mã trạng thái: {response.get('status_code', 'N/A')})")
            else:
                st.success(response.get("message", "Xóa log thành công."))
                st.rerun()
        else:
            st.error("Vui lòng chọn log để xóa.")

    if st.button("Quay lại bước trước"):
        st.session_state.step = st.session_state.previous_step
        st.rerun()

# Giao diện chính
st.title("Ứng dụng AHP - Đánh giá và xếp hạng xe")
st.markdown("Vui lòng thực hiện các bước theo thứ tự hoặc sử dụng file Excel để tính toán.")

with st.sidebar:
    st.header("Điều hướng")
    def change_step():
        st.session_state.previous_step = st.session_state.get("step", "criteria_management")
        st.session_state.step = {
            "Quản lý tiêu chí và xe": "criteria_management",
            "So sánh cặp tiêu chí": "pairwise_comparison",
            "So sánh cặp xe": "alternatives_comparison",
            "Tính AHP từ Excel": "excel_calculation",
            "Xem lịch sử tính toán": "log"
        }[st.session_state.step_select]

    step = st.selectbox(
        "Chọn bước",
        ["Quản lý tiêu chí và xe", "So sánh cặp tiêu chí", "So sánh cặp xe", "Tính AHP từ Excel", "Xem lịch sử tính toán"],
        index=["criteria_management", "pairwise_comparison", "alternatives_comparison", "excel_calculation", "log"].index(st.session_state.get("step", "criteria_management")),
        key="step_select",
        on_change=change_step
    )

if "step" not in st.session_state:
    st.session_state.step = "criteria_management"
if "weights_saved" not in st.session_state:
    st.session_state.weights_saved = False

if st.session_state.step == "criteria_management":
    criteria_management_step()
elif st.session_state.step == "pairwise_comparison":
    pairwise_comparison_step()
elif st.session_state.step == "alternatives_comparison":
    alternatives_comparison_step()
elif st.session_state.step == "excel_calculation":
    excel_calculation_step()
elif st.session_state.step == "log":
    log_step()